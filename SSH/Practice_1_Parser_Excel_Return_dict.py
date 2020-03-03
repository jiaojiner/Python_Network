#!/usr/bin/env python3
# -*- encoding = utf-8 -*-
# 该代码由本人学习时编写，仅供自娱自乐！
# 本人QQ：1945962391
# 欢迎留言讨论，共同学习进步！

from openpyxl import load_workbook


def excel_parser_return_dict(file='test.xlsx', sheel_name='Sheet1'):
    data = load_workbook(file)  # 读取xlsx文件
    table = data[sheel_name]  # 读取sheet数据
    excel_dict = {}
    row_location = 0
    for row in table.iter_rows():
        # print(row)
        if row_location == 0:  # 跳过第一行！
            row_location += 1
            continue
        else:
            cell_location = 0
            for cell in row:
                # print(cell)
                if cell_location == 0:  # 读取第一列的用户名
                    tmp_user = cell.value
                    cell_location += 1
                elif cell_location == 1:  # 读取第二列的密码
                    tmp_pass = cell.value
                    cell_location += 1
                elif cell_location == 2:  # 读取第三列的级别
                    tmp_priv = cell.value
                    cell_location += 1
            excel_dict[tmp_user] = tmp_pass, tmp_priv  # 写入字典
            # print(excel_dict)
        # row_location += 1
        # print(row_location)
    return excel_dict  # 返回字典


if __name__ == "__main__":
    print(excel_parser_return_dict('Practice_1_Read_Accounts.xlsx'))
