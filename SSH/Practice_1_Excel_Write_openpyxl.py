#!/usr/bin/env python3
# -*- encoding = utf-8 -*-
# 该代码由本人学习时编写，仅供自娱自乐！
# 本人QQ：1945962391
# 欢迎留言讨论，共同学习进步！


from openpyxl import Workbook
from openpyxl import load_workbook

dict_excel = {'test123': ('cisco123', 15), 'test456': ('cisco456', 1), 'test789': ('cisco789', 1)}


def excel_write(file='write_pyxl.xlsx', sheel_name='Sheet1', write_dict=dict_excel):
    wb = Workbook()  # 创建xlsx
    ws = wb.create_sheet()  # 创建sheet
    ws.title = sheel_name  # 命名sheet
    # 写入第一行内容
    ws['A1'] = '用户'
    ws['B1'] = '密码'
    ws['C1'] = '级别'
    row_location = 2  # 从第二行开始写入内容
    for x, y in write_dict.items():
        user_locatin = 'A' + str(row_location)
        pass_locatin = 'B' + str(row_location)
        priv_locatin = 'C' + str(row_location)
        ws[user_locatin] = x  # 写入用户
        ws[pass_locatin] = y[0]  # 写入密码
        ws[priv_locatin] = y[1]  # 写入级别
        row_location += 1
    wb.save(file)  # 保存xlsx文件


if __name__ == "__main__":
    excel_write()
